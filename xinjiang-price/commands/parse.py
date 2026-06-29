"""新疆工程造价信息采集 - Excel 解析（同时支持 .xlsx 和 .xls）

数据结构（无论 xlsx/xls）：
- 每个文件多 sheet，每个 sheet 对应一个"县(市)"，sheet 名作为 county 字段
- 每 sheet：
    R0: 标题（合并单元，如"霍尔果斯市2026年4月份建设工程综合价格信息"），也可能是"附件1"
    R1: 表头（合并）—— 序号 | 材料名称及规格型号 | 单位 | 除税综合信息价 | 含税综合信息价
    R2: 分类标题（如"钢材"，可能跨列）—— 也有可能是数据行
    R3+: 数据行 —— [序号, 材料名, 规格, 单位, 不含税, 含税]
- 分类标题行识别：单非空单元 + 在已知分类列表内
- 大表格式（首 sheet 如"伊宁市、伊宁县、察布查尔县"）列数 16000+（合并产生），
  表头是 5 列（序号 | 材料名称及规格型号 | 单位 | 除税综合信息价 | 含税综合信息价）

返回长表 dict：{breed, spec, unit, price, tax_price, category, sheet_name}
"""
import os
import re
import sys
from typing import List, Optional

import openpyxl

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


CATEGORY_HINTS = {
    # 黑色金属
    '钢材', '钢筋', '型钢', '板料', '钢管', '钢丝', '钢绞线', '铁件',
    # 水泥/混凝土
    '水泥', '骨料', '混凝土及其原材料', '混凝土', '砂浆', '外加剂',
    # 砖/砌块/瓦
    '砖、砌块', '砖', '砌块', '瓦',
    # 沥青
    '沥青混凝土', '沥青',
    # 门窗/玻璃
    '门窗', '玻璃',
    # 电
    '电线、电缆', '电缆', '电线', '桥架',
    # 水
    '管材', '管件', '阀门', '法兰',
    # 防水
    '防水材料', '保温材料',
    # 装饰
    '装饰材料', '油漆、涂料', '涂料',
    # 灯具
    '灯具', '开关、插座', '开关插座',
    # 卫生
    '卫生洁具', '水暖器材',
    # 模板
    '模板', '脚手架',
    # 苗木
    '苗木', '园林苗木',
    # 消防/通风
    '消防器材', '通风器材',
}


def _is_category_row(row_vals):
    non_empty = [v for v in row_vals if v is not None and str(v).strip()]
    if len(non_empty) != 1:
        return False
    val = str(non_empty[0]).strip()
    # 支持含空格变体（如 "钢   材" = "钢材"）
    val_normalized = re.sub(r'\s+', '', val)
    hints_normalized = {re.sub(r'\s+', '', h) for h in CATEGORY_HINTS}
    return val in CATEGORY_HINTS or val_normalized in hints_normalized


def _normalize_category(cat):
    """规范化分类名：去多余空格"""
    if not cat:
        return ''
    return re.sub(r'\s+', '', cat).strip()


def _to_str(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_price(s):
    if s is None:
        return None
    s = str(s).strip().replace(',', '').replace(' ', '').replace('￥', '').replace('¥', '')
    if not s:
        return None
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def _find_header_row(vals, start=0):
    # 跳过纯标题/附件行，最多扫 6 行
    for i in range(start, min(start + 6, len(vals))):
        joined = ' '.join(_to_str(v).replace('\n', ' ') for v in vals[i])
        if ('材料名称' in joined and '单位' in joined) or \
           ('序号' in joined and '单位' in joined and ('规格' in joined or '材料' in joined)):
            return i
    return None


def _parse_one_sheet(sheet_rows, sheet_name):
    """解析单个 sheet 的所有行（统一为 List[List[Any]]）。

    sheet_rows: 行数据，每个 row 是 list of cell value
    sheet_name: sheet 名（用于 county 字段）
    """
    if len(sheet_rows) < 3:
        return []

    vals = [[_to_str(c) for c in row] for row in sheet_rows]
    header_idx = _find_header_row(vals, start=0)
    if header_idx is None:
        return []

    headers = [v.replace('\n', ' ').strip() for v in vals[header_idx]]
    n_cols = len(headers)

    # 找价格列
    price_no_tax_col = None
    price_tax_col = None
    for ci, h in enumerate(headers):
        if price_no_tax_col is None and ('除税' in h or '不含税' in h):
            price_no_tax_col = ci
        if price_tax_col is None and '含税' in h:
            price_tax_col = ci

    # 找品种/规格/单位/序号列
    seq_col = breed_col = spec_col = unit_col = None
    for ci, h in enumerate(headers):
        if h == '序号':
            seq_col = ci
        elif '材料名称' in h:
            breed_col = ci
        elif ('规格' in h or '型号' in h) and '材料名称' not in h:
            spec_col = ci
        elif h == '单位':
            unit_col = ci

    # 大表格式（伊犁州首 sheet）：表头只有 1-2 列非空，列数 ≥ 8（合并产生）
    if (sum(1 for h in headers if h) <= 2 or breed_col is None) and n_cols >= 8:
        # 标准 5 列结构：seq, breed（含规格）, unit, 除税, 含税
        seq_col = 0
        breed_col = 1
        spec_col = None
        unit_col = 3
        # 表头常跨两行 → 检查上一行找价格列
        if price_no_tax_col is None or price_tax_col is None:
            for ri in range(max(0, header_idx - 2), header_idx):
                rowh = [v.replace('\n', '').strip() for v in vals[ri]]
                for ci, h in enumerate(rowh):
                    if re.search(r'\d+\s*月\s*除税', h) or re.search(r'除税.*\d+\s*月', h):
                        if price_no_tax_col is None:
                            price_no_tax_col = ci
                    if re.search(r'\d+\s*月\s*含税', h) or re.search(r'含税.*\d+\s*月', h):
                        if price_tax_col is None:
                            price_tax_col = ci
        # 退化：默认 第 4、5 列是价格
        if price_no_tax_col is None:
            price_no_tax_col = 4
        if price_tax_col is None:
            price_tax_col = 5

    if breed_col is None or unit_col is None:
        return []
    if price_no_tax_col is None and price_tax_col is None:
        return []

    out = []
    current_category = ''
    for ri in range(header_idx + 1, len(vals)):
        row_vals = vals[ri]
        if all(v == '' for v in row_vals):
            continue

        if _is_category_row(row_vals):
            current_category = _normalize_category(row_vals[0])
            continue

        def get(col):
            if col is None or col >= len(row_vals):
                return ''
            return row_vals[col]

        seq = get(seq_col)
        breed = get(breed_col)
        spec = get(spec_col) if spec_col is not None else ''
        unit = get(unit_col)

        # 序号列纯数字且 breed 为空：上一行延续（并排双列表）
        if not breed and seq.replace('.', '').isdigit():
            if out:
                last = out[-1]
                breed = last['breed']
                spec = last['spec']
                unit = last['unit']
            else:
                continue

        if not breed and not spec:
            continue

        price_no_tax = _parse_price(get(price_no_tax_col) if price_no_tax_col is not None else '')
        price_tax = _parse_price(get(price_tax_col) if price_tax_col is not None else '')

        if price_no_tax is None and price_tax is None:
            continue

        # 优先用不含税价为 price；只取到含税价时保留
        price = price_no_tax if price_no_tax is not None else price_tax

        out.append({
            'breed': breed,
            'spec': spec,
            'unit': unit,
            'price': price,
            'tax_price': price_tax,
            'category': current_category,
            'sheet_name': sheet_name,
        })
    return out


def _read_xlsx(path):
    """读取 .xlsx → [(sheet_name, [[cell, ...], ...]), ...]"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = []
        for row in ws.iter_rows():
            rows.append([c.value for c in row])
        out.append((sname, rows))
    return out


def _read_xls(path):
    """读取 .xls → [(sheet_name, [[cell, ...], ...]), ...]"""
    if not HAS_XLRD:
        raise RuntimeError('xlrd 未安装，无法读取 .xls 文件。请 pip install xlrd')
    wb = xlrd.open_workbook(path, formatting_info=False)
    out = []
    for sname in wb.sheet_names():
        ws = wb.sheet_by_name(sname)
        rows = []
        for r in range(ws.nrows):
            row = [ws.cell_value(r, c) for c in range(ws.ncols)]
            rows.append(row)
        out.append((sname, rows))
    return out


def parse_excel(path, area_name='', file_basename=''):
    """解析 Excel 文件（.xlsx / .xls 自动判断）→ 长表

    返回 list of dict：{breed, spec, unit, price, tax_price, category, sheet_name}
    """
    low = path.lower()
    if low.endswith('.xlsx'):
        sheets = _read_xlsx(path)
    elif low.endswith('.xls'):
        sheets = _read_xls(path)
    else:
        raise ValueError(f'不支持的文件类型: {path}')

    out = []
    for sname, rows in sheets:
        # 跳过空 sheet
        if not rows or all(all(v is None or str(v).strip() == '' for v in row) for row in rows):
            continue
        out.extend(_parse_one_sheet(rows, sname))
    return out


def parse_xlsx(path, area_name=''):
    """兼容旧接口"""
    return parse_excel(path, area_name)


# ─── CLI 调试 ─────────────────────────────────────────────────────────────
def main():
    import argparse
    parser = argparse.ArgumentParser(description='新疆 Excel 解析（xlsx/xls）')
    parser.add_argument('path', help='Excel 文件路径')
    parser.add_argument('--area', default='', help='地区名（仅显示用）')
    parser.add_argument('--limit', type=int, default=20)
    args = parser.parse_args()

    rows = parse_excel(args.path, args.area)
    print(f'解析 {len(rows)} 条（{os.path.basename(args.path)}）')
    sheets_seen = {}
    for r in rows:
        s = r['sheet_name']
        sheets_seen[s] = sheets_seen.get(s, 0) + 1
    print('各 sheet 条数：')
    for s, c in sheets_seen.items():
        print(f'  {s}: {c}')
    print()
    print('样本（前 {} 条）：'.format(args.limit))
    for r in rows[:args.limit]:
        print(' ', r)


if __name__ == '__main__':
    main()


# （上面的 main() 已保留）


# ─── breed / spec 拆分 ───────────────────────────────────────────────────────
# 新疆源数据的"材料名称及规格型号"是把品种和规格写在同一列的：
#   "低碳热轧盘条（高线）HPB300 Φ6"  →  breed="低碳热轧盘条（高线）"  spec="HPB300 Φ6"
#   "商品混凝土C20 （泵送）（二级配）" →  breed="商品混凝土"           spec="C20 （泵送）（二级配）"
# 用 SPEC_PATTERNS 识别规格特征（从右往左找最左侧的特征为切点）。

SPEC_PATTERNS = [
    # 多特征组合（先匹配）
    (r'HRB\d+\w*\s+Φ\d+(?:\.\d+)?(?:以上)?', 'steel_grade+phi'),
    (r'HPB\d+\w*\s+Φ\d+(?:\.\d+)?(?:以上)?', 'steel_grade+phi'),
    (r'Q\d+\w*\s+\d+(?:\.\d+)?(?:×\d+(?:\.\d+)?)*', 'grade+size'),
    # 沥青
    (r'AC-\d+', 'asphalt'),
    # PPR 管
    (r'De\d+(?:\.\d+)?×\d+(?:\.\d+)?', 'ppr_size'),
    # 水泥标号 P.O42.5R
    (r'P\.[OS]\d+\.?\d*[RS]?', 'cement_grade'),
    # 混凝土强度 + 后续括号描述（吃到底）
    (r'C\d{2,3}(?:\s*[（(][^)）]*[)）])*', 'concrete_grade_full'),
    # 公称直径 + 后续括号
    (r'DN\d+(?:\([^)]+\))?', 'dn'),
    # 直径
    (r'Φ\d+(?:\.\d+)?(?:以上)?', 'phi'),
    # 电线电缆规格：NH-BV-10 / ZRBV-25 / BVR-4 / YJV22-4×6 / BYJ(F)-10 / YJV 3*185+2*95
    (r'\b(?:NH|ZR|WD|WDZ|WDZB|WDZN|ZR-|ZC|N|ZB|YJV|BV|BVR|BYJ|BYJ\(F\)|RVV|RVS|RVP|RVVP|RVSP|KVV|KVV22|YJV22|YJLV22)[-\w()]*\s*\d+[×*+]\d+(?:\.\d+)?(?:[×*+]\d+(?:\.\d+)?)*', 'cable_multi'),
    (r'\b(?:NH|ZR|WD|WDZ|WDZB|WDZN|ZR-|ZC|N|ZB|YJV|BV|BVR|BYJ|BYJ\(F\)|RVV|RVS|RVP|RVVP|RVSP|KVV|KVV22|YJV22|YJLV22)[-\w()]*\d+(?:\.\d+)?(?:[×*]\d+(?:\.\d+)?)*', 'cable'),
    (r'BVR?\d+(?:\.\d+)?', 'wire'),
    # 钢牌号 + 尺寸
    (r'Q\d+\w*\s+\d+(?:\.\d+)?', 'grade+num'),
    # 工字钢/槽钢型号：I20a、I25a、[10、[16a
    (r'(?:^|[\s\u3000（(])[I\[]\d+[a-zA-Z]?(?=[\s\u3000）)\u3001、，,]|$)', 'steel_section'),
    # H 型钢/角钢截面尺寸：100*100*6*8 / 100×100 / 75*50*6
    (r'\d+(?:\.\d+)?[*×]\d+(?:\.\d+)?(?:[*×]\d+(?:\.\d+)?)*', 'h_section_dim'),
    # 直埋式保温管、无缝钢管：D100、D219（区分 De\d+ PPR管的 "De"）
    (r'(?:^|[\s\u3000（(])D\d+(?:\.\d+)?(?=[\s\u3000）)\u3001、，,直]|$)', 'd_pipe_size'),
    # 钢塑复合压力管 / PPR：De110、De110×10
    (r'De\d+(?:\.\d+)?(?=[\s\u3000）)\u3001、，,]|$)', 'de_dn'),
    # SYV 视频线：SYV75-5、SYV75-7
    (r'(?:^|[\s\u3000（(])SYV\d+(?:-\d+)?(?=[^a-zA-Z\d]|$)', 'syv_cable'),
    # UTP 网络线：UTP-6
    (r'(?:^|[\s\u3000（(])UTP-\d+(?=[^a-zA-Z\d]|$)', 'utp_cable'),
    # 铜芯橡皮绝缘线：BX4、BX6
    (r'(?:^|[\s\u3000（(])BX\d+(?:\.\d+)?(?=[^a-zA-Z\d]|$)', 'bx_cable'),
    # 铝芯橡皮绝缘线：BLX4、BLX120
    (r'(?:^|[\s\u3000（(])BLX\d+(?:\.\d+)?(?=[^a-zA-Z\d]|$)', 'blx_cable'),
    # 铝芯塑料绝缘线：BLV4、BLV6
    (r'(?:^|[\s\u3000（(])BLV\d+(?:\.\d+)?(?=[^a-zA-Z\d]|$)', 'blv_cable'),
    # 凿散热器型号：GZ_2、GZ-3
    (r'GZ[_-]?\d+(?=[\s\u3000、，,.．）)）\u3001]|$)', 'radiator_model'),
    # 砂浆强度：M5、M10、M15
    (r'(?:^|[\s\u3000（(])M\d+(?:[.．\u3001、，,]|$|\s)', 'mortar_grade'),
    # 地暖分水器 / 回路数：2路、3路
    (r'(?:^|[\s\u3000（(])\d+路(?=[\s\u3000）)\u3001、，,]|$)', 'way_count'),
    # 电气规格：10A 250V
    (r'\d+\.?\d*\s*[AV](?:\s+\d+\.?\d*\s*[AV])', 'voltage_current'),
    # 燃油标号：92# / 95# / 0# / -10# / -20# / -35#
    (r'-?\d+(?:\.\d+)?#(?![\u4e00-\u9fa5])', 'fuel_grade'),
    # 铝合金/塑钢门窗系列：65系列 / 70系列 / 75系列
    (r'\d+系列', 'window_series'),
    (r'Q\d+\w+', 'grade'),
    (r'HRB\d+\w*', 'hrb'),
    (r'HPB\d+\w*', 'hpb'),
    # 带单位数值
    (r'\d+\.?\d*\s*(?:mm|cm|m|MPa|MPa|kg|级)', 'unit_num'),
    # 小数（水泥强度 42.5）
    (r'\d+\.\d+(?:[RS])?', 'decimal'),
]


def split_breed_spec(breed: str) -> tuple[str, str]:
    """把"材料名+规格"拆成 (breed_clean, spec)。

    规则：
    - 从右向左扫描 SPEC_PATTERNS，找到第一个规格特征，从那里切分
    - 混凝土强度 + 括号描述（吃到底）
    - DN 后续括号吃掉

    返回: (cleaned_breed, spec) 或 (original_breed, '') 如果无法拆分
    """
    if not breed:
        return '', ''
    s = str(breed).strip()
    if not s:
        return '', ''

    # 选最早匹配的特征作为切分点
    best = None
    for pat, kind in SPEC_PATTERNS:
        for m in re.finditer(pat, s):
            if best is None or m.start() < best[0]:
                best = (m.start(), m.end(), m.group(0).strip(), kind)
        if best:
            break

    if best is None:
        return s, ''

    start, end, spec_val, kind = best

    # 向前回退：如果 spec 前是空格，去掉空格
    breed_end = start
    if start > 0 and s[start - 1] in ' \u3000':
        breed_end = start - 1

    # spec 部分扩展（部分模式吃后续括号）
    spec_end = end
    if kind == 'concrete_grade_full':
        # C20 后面吃所有 "（...）" 块
        pos = spec_end
        while pos < len(s):
            # 跳过空格
            while pos < len(s) and s[pos] in ' \u3000':
                pos += 1
            if pos < len(s) and s[pos] in '（(':
                depth = 0
                i = pos
                while i < len(s):
                    ch = s[i]
                    if ch in '（(':
                        depth += 1
                    elif ch in '）)':
                        depth -= 1
                        if depth == 0:
                            spec_end = i + 1
                            pos = spec_end
                            break
                    i += 1
                else:
                    break
            else:
                break
    elif kind == 'dn' and end < len(s) and s[end] == '(':
        depth = 0
        for i in range(end, len(s)):
            ch = s[i]
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
                if depth == 0:
                    spec_end = i + 1
                    break

    breed_part = s[:breed_end].rstrip().rstrip('，,').rstrip()
    spec_part = s[start:spec_end].strip()

    # 清理：breed 末尾如果是未配对的"（"，去掉
    if breed_part.endswith('（') and not spec_part.startswith('）'):
        breed_part = breed_part[:-1].rstrip()
    elif breed_part.endswith('(') and not spec_part.startswith(')'):
        breed_part = breed_part[:-1].rstrip()

    # 兜底：breed 拆完为空
    if not breed_part:
        # spec 在开头（如"65系列单框三玻..."）→ spec 后剩余部分作为 breed
        remainder = s[spec_end:].strip()
        if remainder and len(remainder) >= 2:
            breed_part = remainder
            return breed_part, spec_part
        return s, ''

    return breed_part, spec_part
